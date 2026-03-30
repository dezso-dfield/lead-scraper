from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scraper.models import Lead


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
HYPERLINK_FONT = Font(color="0563C1", underline="single")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def export_excel(leads: list[Lead], path: str | Path, query: str = "", location: str = "") -> Path:
    path = Path(path)
    sorted_leads = sorted(leads, key=lambda l: l.confidence, reverse=True)

    rows = [l.to_dict() for l in sorted_leads]
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)

        # Summary sheet
        summary_data = {
            "Metric": ["Query", "Location", "Total Leads", "With Email", "With Phone", "With Both", "Average Confidence"],
            "Value": [
                query,
                location,
                len(leads),
                sum(1 for l in leads if l.emails),
                sum(1 for l in leads if l.phones),
                sum(1 for l in leads if l.emails and l.phones),
                f"{sum(l.confidence for l in leads) / max(len(leads), 1):.0%}",
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

    # Apply styling
    wb = load_workbook(path)
    _style_leads_sheet(wb["Leads"])
    _style_summary_sheet(wb["Summary"])
    wb.save(path)

    return path


def _style_leads_sheet(ws):
    col_widths = {
        "A": 35,  # company_name
        "B": 40,  # website
        "C": 45,  # emails
        "D": 22,  # phones
        "E": 40,  # address
        "F": 20,  # city
        "G": 15,  # country
        "H": 30,  # niche
        "I": 25,  # sources
        "J": 12,  # confidence
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Style header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Style data rows + add hyperlinks
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = BORDER

        # Hyperlink website (column B = index 1)
        website_cell = row[1]
        val = website_cell.value or ""
        if val.startswith("http"):
            website_cell.hyperlink = val
            website_cell.font = HYPERLINK_FONT

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _style_summary_sheet(ws):
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER
        row[0].font = Font(bold=True)
